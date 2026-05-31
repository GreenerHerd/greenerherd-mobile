#!/usr/bin/env python3
"""Extract nutrition requirement profiles from Livestock Nutrition Requirements.xlsx."""

from __future__ import annotations

import argparse
import json
import re
import uuid
from pathlib import Path

import openpyxl

SCRIPT_DIR = Path(__file__).resolve().parent
DB_DIR = SCRIPT_DIR.parent
OUT_DIR = DB_DIR / 'seeds'
SOURCES_DIR = DB_DIR / 'sources'
MOBILE_ASSETS_DIR = DB_DIR.parent.parent / 'assets' / 'data'
DEFAULT_XLSX = SOURCES_DIR / 'Livestock_Nutrition_Requirements.xlsx'
FALLBACK_XLSX = (
    Path.home()
    / 'Documents/GreenerHerd/Copy of Livestock Nutrition Requirements.xlsx'
)

NS = uuid.UUID('8f4e3c21-9b2a-4f6d-a1e0-000000000002')


def slug(*parts: str) -> str:
    s = '_'.join(p for p in parts if p)
    s = re.sub(r'[^a-zA-Z0-9]+', '_', s.strip())
    return re.sub(r'_+', '_', s).strip('_').upper()


def sql_str(v: object | None) -> str:
    if v is None:
        return 'NULL'
    return "'" + str(v).replace("'", "''") + "'"


def sql_num(v: object | None) -> str:
    if v is None or v == '':
        return 'NULL'
    try:
        return str(round(float(v), 6))
    except (TypeError, ValueError):
        return 'NULL'


def sql_int(v: object | None) -> str:
    if v is None or v == '':
        return 'NULL'
    try:
        return str(int(float(v)))
    except (TypeError, ValueError):
        return 'NULL'


def fnum(v: object | None) -> float | None:
    if v is None or v == '':
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def profile_id(code: str) -> str:
    return str(uuid.uuid5(NS, code))


def compute_feed_cycle(p: dict) -> str:
    """Standard feed-cycle code for UI / animal matching (reproductive & growth phase)."""
    ps = p['production_system']
    ls = p['life_stage']
    m = p.get('months_since_calving')

    if ps == 'DAIRY':
        dairy = {
            'Dry (Far off)': 'DRY_FAR_OFF',
            'Close (Close up)': 'CLOSE_UP',
            'Cow - Fresh': 'FRESH',
            'Cow - Early': 'EARLY_LACTATION',
            'Cow - Mid': 'MID_LACTATION',
            'Cow - Late': 'LATE_LACTATION',
            '6-mo Heifer': 'HEIFER_6MO',
            '12-mo Heifer': 'HEIFER_12MO',
            '18-mo Heifer': 'HEIFER_18MO',
            '24-mo Heifer (Close up)': 'HEIFER_24MO_CLOSE_UP',
        }
        return dairy.get(ls, slug(ls))

    if ps == 'SMALL_RUMINANT':
        if ls == 'Sick':
            return 'SICK'
        if ls == 'Weaning':
            return 'WEANING'
        if ls == 'Breeding':
            return 'BREEDING'
        ac = p.get('animal_class') or ''
        base = {
            'Maintenance': 'MAINTENANCE',
            'Lactating': 'LACTATING',
            'Late pregnancy': 'PREGNANT',
            'Fattening': 'FATTENING',
        }.get(ls, 'MAINTENANCE')
        if ac == 'Young':
            return f'{base}_YOUNG'
        return base

    if ps == 'DAIRY' and ls == 'Breeding bull':
        return 'BREEDING'

    if ps == 'BEEF' and ls == 'Maintenance':
        return 'MAINTENANCE'
    if ps == 'BEEF' and ls == 'Sick':
        return 'SICK'
    if ps == 'BEEF' and ls == 'Weaning':
        return 'WEANING'
    if ls == 'Growing':
        return 'GROWING'
    stage = {
        'Early Lactation': 'EARLY_LACTATION',
        'Mid Lactation': 'MID_LACTATION',
        'Mid Gestation': 'MID_GESTATION',
        'Late Gestation': 'LATE_GESTATION',
    }.get(ls, slug(ls).upper())
    if m is not None:
        return f'{stage}_M{int(m)}'
    return stage


def attach_feed_cycle(profiles: list[dict]) -> None:
    for p in profiles:
        p['feed_cycle'] = compute_feed_cycle(p)


def parse_dairy(wb: openpyxl.Workbook) -> list[dict]:
    ws = wb['Dairy Cattle']
    rows = list(ws.iter_rows(values_only=True))
    stages = [str(s).strip() for s in rows[0][1:] if s]
    metrics: dict[str, list] = {}
    for r in rows[1:30]:
        if r[0]:
            metrics[str(r[0]).strip()] = list(r[1:])

    def metric_val(name: str, idx: int) -> float | None:
        for key, vals in metrics.items():
            if name in key or key.startswith(name):
                if idx < len(vals):
                    return fnum(vals[idx])
        return None

    profiles: list[dict] = []
    for i, stage in enumerate(stages):
        dmi = metric_val('DMI', i)
        if not dmi:
            continue
        code = f'CATTLE_DAIRY_{slug(stage)}'
        cp = metric_val('Crude protein', i)
        nel = metric_val('NEL', i)
        me = metric_val('ME', i)
        energy_per_kg = nel if nel is not None else me
        ndf = metric_val('NDF', i)
        adf = metric_val('ADF', i)
        ca = metric_val('Calcium', i)
        p = metric_val('Phosphorus', i)
        body = metric_val('Body wt', i)
        milk = metric_val('Milk production', i)
        profiles.append(
            {
                'id': profile_id(code),
                'profile_code': code,
                'species': 'CATTLE',
                'production_system': 'DAIRY',
                'life_stage': stage,
                'animal_class': None,
                'body_weight_kg': body,
                'months_since_calving': None,
                'dmi_kg_day': dmi,
                'cp_percent_dm': cp,
                'nel_mcal_per_kg_dm': energy_per_kg,
                'ndf_percent_dm': ndf,
                'adf_percent_dm': adf,
                'ca_percent_dm': ca,
                'p_percent_dm': p,
                'milk_kg_day': milk,
                'tdn_kg_day': None,
                'nem_mcal_day': energy_per_kg * dmi if energy_per_kg and dmi else None,
                'neg_mcal_day': None,
                'cp_kg_day': dmi * cp / 100 if cp else None,
                'ca_kg_day': dmi * ca / 100 if ca else None,
                'p_kg_day': dmi * p / 100 if p else None,
                'fodder_percent': 50,
                'concentrate_percent': 40,
                'source_sheet': 'Dairy Cattle',
            }
        )
    return profiles


LB_TO_KG = 0.453592

CALENDAR_STAGES = frozenset(
    {
        'Early Lactation',
        'Mid Lactation',
        'Mid Gestation',
        'Late Gestation',
    }
)


def _row_len(r: tuple) -> list:
    return list(r) if r else []


def _is_beef_section_header(r: tuple) -> bool:
    if not r or not r[0] or not isinstance(r[0], str):
        return False
    label = str(r[0]).strip()
    if label.startswith('Source') or 'http' in label:
        return False
    col5 = str(r[5]).strip() if len(r) > 5 and r[5] is not None else ''
    col1 = str(r[1]).strip() if len(r) > 1 and r[1] is not None else ''
    if col5 == 'Stage' or 'Stage' in col5:
        return True
    if 'ADG' in col5 or 'Weight Range' in col5:
        return True
    if 'Avg Weight' in col1 or 'Mature Weight' in col1:
        return True
    return False


def _intake_dm_kg(row: list, layout: str) -> float | None:
    """Calendar sections use kg (col 18); growth sections often use lbs (col 19)."""
    v18 = fnum(row[18]) if len(row) > 18 else None
    v19 = fnum(row[19]) if len(row) > 19 else None
    if layout == 'calendar':
        if v18 is not None and v18 >= 0.5:
            return v18
        if v19 is not None and v19 >= 0.5:
            return v19 * LB_TO_KG
        return None
    if v19 is not None and v19 >= 0.5:
        return v19 * LB_TO_KG
    if v18 is not None and v18 >= 0.5:
        return v18
    return None


def parse_beef(wb: openpyxl.Workbook) -> list[dict]:
    ws = wb['Beef Cattle kg)']
    rows = list(ws.iter_rows(values_only=True))
    profiles: list[dict] = []
    current_class: str | None = None
    layout: str = 'calendar'

    for r in rows:
        row = _row_len(r)
        if not row or all(x is None for x in row):
            continue

        if _is_beef_section_header(tuple(row)):
            current_class = str(row[0]).strip()
            col5 = str(row[5]).strip() if len(row) > 5 and row[5] is not None else ''
            layout = 'growth' if 'ADG' in col5 or 'Weight Range' in col5 else 'calendar'
            continue

        if not current_class:
            continue

        stage_raw = row[5] if len(row) > 5 else None
        stage = str(stage_raw).strip() if stage_raw is not None else ''
        if not stage or stage in ('Stage', '#VALUE!'):
            continue

        intake_dm = _intake_dm_kg(row, layout)
        if intake_dm is None or intake_dm < 0.3:
            continue

        months = fnum(row[3]) if len(row) > 3 else None
        weight = fnum(row[2]) if len(row) > 2 else None
        if weight is not None and weight > 400:
            weight = weight * LB_TO_KG

        tdn_kg = fnum(row[13]) if len(row) > 13 else None
        cp_kg = fnum(row[15]) if len(row) > 15 else None
        ca_kg = fnum(row[16]) if len(row) > 16 else None
        p_kg = fnum(row[17]) if len(row) > 17 else None
        nem_mcal = fnum(row[8]) if len(row) > 8 else None

        if layout == 'calendar':
            if stage not in CALENDAR_STAGES:
                continue
            life_stage = stage
            months_int = int(months) if months is not None else None
            code = f'CATTLE_BEEF_{slug(current_class)}_{slug(life_stage)}'
            if months_int is not None:
                code = f'{code}_M{months_int}'
        else:
            life_stage = 'Growing'
            weight_lb = fnum(row[3]) if len(row) > 3 else None
            adg = fnum(row[5]) if len(row) > 5 else None
            code = f'CATTLE_BEEF_{slug(current_class)}_GROWING'
            if weight_lb is not None:
                code = f'{code}_W{int(weight_lb)}'
            elif adg is not None:
                code = f'{code}_ADG{str(adg).replace(".", "_")}'
            months_int = None

        profiles.append(
            {
                'id': profile_id(code),
                'profile_code': code,
                'species': 'CATTLE',
                'production_system': 'BEEF',
                'life_stage': life_stage,
                'animal_class': current_class,
                'body_weight_kg': weight,
                'months_since_calving': months_int,
                'dmi_kg_day': intake_dm,
                'cp_percent_dm': fnum(row[24]) if len(row) > 24 else fnum(row[22]),
                'nel_mcal_per_kg_dm': None,
                'ndf_percent_dm': None,
                'adf_percent_dm': None,
                'ca_percent_dm': fnum(row[25]) if len(row) > 25 else fnum(row[23]),
                'p_percent_dm': fnum(row[26]) if len(row) > 26 else fnum(row[24]),
                'milk_kg_day': None,
                'tdn_kg_day': tdn_kg,
                'nem_mcal_day': nem_mcal,
                'neg_mcal_day': None,
                'cp_kg_day': cp_kg,
                'ca_kg_day': ca_kg,
                'p_kg_day': p_kg,
                'fodder_percent': 50,
                'concentrate_percent': 40,
                'source_sheet': 'Beef Cattle kg)',
            }
        )
    return profiles


def _scaled_row(
    *,
    code: str,
    species: str,
    production_system: str,
    life_stage: str,
    feed_cycle: str,
    dmi_kg_day: float,
    animal_class: str | None = None,
    cp_percent_dm: float | None = None,
    cp_kg_day: float | None = None,
    tdn_kg_day: float | None = None,
    nem_mcal_day: float | None = None,
    nel_mcal_per_kg_dm: float | None = None,
    ca_kg_day: float | None = None,
    p_kg_day: float | None = None,
    body_weight_kg: float | None = None,
    ndf_percent_dm: float | None = None,
    adf_percent_dm: float | None = None,
    ca_percent_dm: float | None = None,
    p_percent_dm: float | None = None,
    fodder_percent: float = 50,
    concentrate_percent: float = 40,
) -> dict:
    return {
        'id': profile_id(code),
        'profile_code': code,
        'species': species,
        'production_system': production_system,
        'life_stage': life_stage,
        'animal_class': animal_class,
        'body_weight_kg': body_weight_kg,
        'months_since_calving': None,
        'dmi_kg_day': round(dmi_kg_day, 6),
        'cp_percent_dm': cp_percent_dm,
        'nel_mcal_per_kg_dm': nel_mcal_per_kg_dm,
        'ndf_percent_dm': ndf_percent_dm,
        'adf_percent_dm': adf_percent_dm,
        'ca_percent_dm': ca_percent_dm,
        'p_percent_dm': p_percent_dm,
        'milk_kg_day': None,
        'tdn_kg_day': tdn_kg_day,
        'nem_mcal_day': nem_mcal_day,
        'neg_mcal_day': None,
        'cp_kg_day': cp_kg_day,
        'ca_kg_day': ca_kg_day,
        'p_kg_day': p_kg_day,
        'fodder_percent': fodder_percent,
        'concentrate_percent': concentrate_percent,
        'source_sheet': 'derived_extension',
        'feed_cycle': feed_cycle,
    }


def _find_profile(profiles: list[dict], **match) -> dict | None:
    for p in profiles:
        if all(p.get(k) == v for k, v in match.items()):
            return p
    return None


def derived_extension_profiles(profiles: list[dict]) -> list[dict]:
    """Sick, weaning, dairy breeding bull, beef maintenance, young/adult small ruminants."""
    by_code = {p['profile_code']: p for p in profiles}
    out: list[dict] = []

    mature_cow_m7 = by_code.get('CATTLE_BEEF_MATURE_COW_MID_GESTATION_M7')
    first_heifer_m7 = by_code.get('CATTLE_BEEF_FIRST_CALF_HEIFER_MID_GESTATION_M7')

    if mature_cow_m7:
        out.append(
            _scaled_row(
                code='CATTLE_BEEF_MATURE_COW_MAINTENANCE',
                species='CATTLE',
                production_system='BEEF',
                life_stage='Maintenance',
                feed_cycle='MAINTENANCE',
                animal_class='Mature Cow',
                dmi_kg_day=mature_cow_m7['dmi_kg_day'],
                cp_kg_day=mature_cow_m7.get('cp_kg_day'),
                tdn_kg_day=mature_cow_m7.get('tdn_kg_day'),
                nem_mcal_day=mature_cow_m7.get('nem_mcal_day'),
                ca_kg_day=mature_cow_m7.get('ca_kg_day'),
                p_kg_day=mature_cow_m7.get('p_kg_day'),
                body_weight_kg=mature_cow_m7.get('body_weight_kg'),
            )
        )
        out.append(
            _scaled_row(
                code='CATTLE_SICK',
                species='CATTLE',
                production_system='BEEF',
                life_stage='Sick',
                feed_cycle='SICK',
                animal_class='Mature Cow',
                dmi_kg_day=mature_cow_m7['dmi_kg_day'] * 0.85,
                cp_kg_day=(mature_cow_m7.get('cp_kg_day') or 0) * 0.9,
                tdn_kg_day=(mature_cow_m7.get('tdn_kg_day') or 0) * 0.85,
                nem_mcal_day=(mature_cow_m7.get('nem_mcal_day') or 0) * 0.85,
                ca_kg_day=(mature_cow_m7.get('ca_kg_day') or 0) * 0.9,
                p_kg_day=(mature_cow_m7.get('p_kg_day') or 0) * 0.9,
                body_weight_kg=mature_cow_m7.get('body_weight_kg'),
            )
        )

    if first_heifer_m7:
        out.append(
            _scaled_row(
                code='CATTLE_BEEF_FIRST_CALF_HEIFER_MAINTENANCE',
                species='CATTLE',
                production_system='BEEF',
                life_stage='Maintenance',
                feed_cycle='MAINTENANCE',
                animal_class='First Calf Heifer',
                dmi_kg_day=first_heifer_m7['dmi_kg_day'],
                cp_kg_day=first_heifer_m7.get('cp_kg_day'),
                tdn_kg_day=first_heifer_m7.get('tdn_kg_day'),
                nem_mcal_day=first_heifer_m7.get('nem_mcal_day'),
                ca_kg_day=first_heifer_m7.get('ca_kg_day'),
                p_kg_day=first_heifer_m7.get('p_kg_day'),
                body_weight_kg=first_heifer_m7.get('body_weight_kg'),
            )
        )

    out.append(
        _scaled_row(
            code='CATTLE_BEEF_MATURE_BULL_MAINTENANCE',
            species='CATTLE',
            production_system='BEEF',
            life_stage='Maintenance',
            feed_cycle='MAINTENANCE',
            animal_class='Mature Bull Maintenance',
            dmi_kg_day=11.5,
            cp_kg_day=0.92,
            tdn_kg_day=6.9,
            nem_mcal_day=9.2,
            ca_kg_day=0.028,
            p_kg_day=0.020,
            body_weight_kg=900,
        )
    )

    out.append(
        _scaled_row(
            code='CATTLE_DAIRY_BREEDING_BULL',
            species='CATTLE',
            production_system='DAIRY',
            life_stage='Breeding bull',
            feed_cycle='BREEDING',
            animal_class='Breeding bull',
            dmi_kg_day=12.0,
            cp_percent_dm=12.0,
            nel_mcal_per_kg_dm=2.0,
            ndf_percent_dm=32.0,
            adf_percent_dm=22.0,
            ca_percent_dm=0.45,
            p_percent_dm=0.30,
            nem_mcal_day=24.0,
            cp_kg_day=1.44,
            ca_kg_day=0.054,
            p_kg_day=0.036,
            body_weight_kg=950,
        )
    )

    calf_growing = _find_profile(
        profiles,
        production_system='BEEF',
        animal_class='Feeder Calves <11 months of age',
        life_stage='Growing',
    )
    if calf_growing:
        out.append(
            _scaled_row(
                code='CATTLE_WEANING',
                species='CATTLE',
                production_system='BEEF',
                life_stage='Weaning',
                feed_cycle='WEANING',
                animal_class='Feeder Calves <11 months of age',
                dmi_kg_day=calf_growing['dmi_kg_day'] * 0.9,
                cp_kg_day=(calf_growing.get('cp_kg_day') or 0) * 1.05,
                tdn_kg_day=(calf_growing.get('tdn_kg_day') or 0) * 0.9,
                nem_mcal_day=(calf_growing.get('nem_mcal_day') or 0) * 0.9,
                ca_kg_day=(calf_growing.get('ca_kg_day') or 0) * 1.05,
                p_kg_day=(calf_growing.get('p_kg_day') or 0) * 1.05,
                body_weight_kg=(calf_growing.get('body_weight_kg') or 0) * 0.85,
            )
        )

    sr_phases = [
        ('MAINTENANCE', 'Maintenance', 'SMALL_RUMINANT_MAINTENANCE'),
        ('LACTATING', 'Lactating', 'SMALL_RUMINANT_LACTATING'),
        ('PREGNANT', 'Late pregnancy', 'SMALL_RUMINANT_PREGNANT'),
        ('FATTENING', 'Fattening', 'SMALL_RUMINANT_FATTENING'),
    ]
    for species in ('GOAT', 'SHEEP'):
        for feed_cycle, life_stage, suffix in sr_phases:
            adult_code = f'{species}_{suffix}'
            adult = by_code.get(adult_code)
            if not adult:
                continue

            young_dmi = adult['dmi_kg_day'] * 0.65
            cp_kg = (adult.get('cp_kg_day') or 0) * 0.75
            out.append(
                _scaled_row(
                    code=f'{species}_{suffix}_YOUNG',
                    species=species,
                    production_system='SMALL_RUMINANT',
                    life_stage=life_stage,
                    feed_cycle=f'{feed_cycle}_YOUNG',
                    animal_class='Young',
                    dmi_kg_day=young_dmi,
                    cp_kg_day=cp_kg,
                    tdn_kg_day=(adult.get('tdn_kg_day') or 0) * 0.65,
                    ca_kg_day=(adult.get('ca_kg_day') or 0) * 0.75,
                    p_kg_day=(adult.get('p_kg_day') or 0) * 0.75,
                    body_weight_kg=(adult.get('body_weight_kg') or 0) * 0.55,
                    cp_percent_dm=(adult.get('cp_percent_dm') or 0) * 1.05
                    if adult.get('cp_percent_dm')
                    else None,
                    fodder_percent=adult.get('fodder_percent') or 55,
                    concentrate_percent=adult.get('concentrate_percent') or 35,
                )
            )

        maint = by_code.get(f'{species}_SMALL_RUMINANT_MAINTENANCE')
        if maint:
            out.append(
                _scaled_row(
                    code=f'{species}_SICK',
                    species=species,
                    production_system='SMALL_RUMINANT',
                    life_stage='Sick',
                    feed_cycle='SICK',
                    animal_class='Adult',
                    dmi_kg_day=maint['dmi_kg_day'] * 0.85,
                    cp_kg_day=(maint.get('cp_kg_day') or 0) * 0.9,
                    tdn_kg_day=(maint.get('tdn_kg_day') or 0) * 0.85,
                    ca_kg_day=(maint.get('ca_kg_day') or 0) * 0.9,
                    p_kg_day=(maint.get('p_kg_day') or 0) * 0.9,
                    body_weight_kg=maint.get('body_weight_kg'),
                )
            )
            out.append(
                _scaled_row(
                    code=f'{species}_SMALL_RUMINANT_WEANING',
                    species=species,
                    production_system='SMALL_RUMINANT',
                    life_stage='Weaning',
                    feed_cycle='WEANING',
                    animal_class='Young',
                    dmi_kg_day=maint['dmi_kg_day'] * 0.55,
                    cp_kg_day=(maint.get('cp_kg_day') or 0) * 0.85,
                    tdn_kg_day=(maint.get('tdn_kg_day') or 0) * 0.55,
                    ca_kg_day=(maint.get('ca_kg_day') or 0) * 0.85,
                    p_kg_day=(maint.get('p_kg_day') or 0) * 0.85,
                    body_weight_kg=(maint.get('body_weight_kg') or 0) * 0.35,
                    cp_percent_dm=14.0,
                    fodder_percent=60,
                    concentrate_percent=30,
                )
            )
            out.append(
                _scaled_row(
                    code=f'{species}_SMALL_RUMINANT_BREEDING',
                    species=species,
                    production_system='SMALL_RUMINANT',
                    life_stage='Breeding',
                    feed_cycle='BREEDING',
                    animal_class='Adult',
                    dmi_kg_day=maint['dmi_kg_day'] * 1.05,
                    cp_kg_day=(maint.get('cp_kg_day') or 0) * 1.08,
                    tdn_kg_day=(maint.get('tdn_kg_day') or 0) * 1.05,
                    ca_kg_day=(maint.get('ca_kg_day') or 0) * 1.05,
                    p_kg_day=(maint.get('p_kg_day') or 0) * 1.05,
                    body_weight_kg=maint.get('body_weight_kg'),
                )
            )

    return out


def small_ruminant_profiles() -> list[dict]:
    """Reference profiles (not in xlsx) for sheep/goats — NRC-style defaults."""
    # (suffix, stage, weight kg, DMI kg/d, CP fraction of DMI, TDN kg/d, Ca g/d, P g/d, fodder %, conc %)
    base = [
        (
            'SMALL_RUMINANT_MAINTENANCE',
            'Maintenance',
            45,
            1.5,
            0.12,
            0.9,
            4.0,
            3.5,
            50,
            35,
        ),
        (
            'SMALL_RUMINANT_LACTATING',
            'Lactating',
            55,
            2.8,
            0.22,
            1.6,
            7.5,
            5.5,
            55,
            40,
        ),
        (
            'SMALL_RUMINANT_PREGNANT',
            'Late pregnancy',
            60,
            2.2,
            0.16,
            1.2,
            6.0,
            4.5,
            55,
            38,
        ),
        (
            'SMALL_RUMINANT_FATTENING',
            'Fattening',
            40,
            2.5,
            0.18,
            1.4,
            6.5,
            4.0,
            45,
            50,
        ),
    ]
    profiles: list[dict] = []
    for species in ('GOAT', 'SHEEP'):
        for code_suffix, stage, wt, dmi, cp_frac, tdn_kg, ca_g, p_g, fod, conc in base:
            code = f'{species}_{code_suffix}'
            profiles.append(
                {
                    'id': profile_id(code),
                    'profile_code': code,
                    'species': species,
                    'production_system': 'SMALL_RUMINANT',
                    'life_stage': stage,
                    'animal_class': None,
                    'body_weight_kg': wt,
                    'months_since_calving': None,
                    'dmi_kg_day': dmi,
                    'cp_percent_dm': cp_frac * 100 if dmi else None,
                    'nel_mcal_per_kg_dm': None,
                    'ndf_percent_dm': None,
                    'adf_percent_dm': None,
                    'ca_percent_dm': None,
                    'p_percent_dm': None,
                    'milk_kg_day': None,
                    'tdn_kg_day': tdn_kg,
                    'nem_mcal_day': None,
                    'neg_mcal_day': None,
                    'cp_kg_day': dmi * cp_frac,
                    'ca_kg_day': ca_g / 1000,
                    'p_kg_day': p_g / 1000,
                    'fodder_percent': fod,
                    'concentrate_percent': conc,
                    'source_sheet': 'derived',
                }
            )
    return profiles


def write_sql(profiles: list[dict], path: Path) -> None:
    lines = [
        '-- Nutrition requirement profiles (from Livestock Nutrition Requirements.xlsx)',
        'TRUNCATE nutrition_requirement_profiles RESTART IDENTITY CASCADE;',
        '',
    ]
    for p in profiles:
        lines.append(
            f"""INSERT INTO nutrition_requirement_profiles (
  id, profile_code, species, production_system, life_stage, animal_class, feed_cycle,
  body_weight_kg, months_since_calving, dmi_kg_day,
  cp_percent_dm, nel_mcal_per_kg_dm, ndf_percent_dm, adf_percent_dm,
  ca_percent_dm, p_percent_dm, milk_kg_day,
  tdn_kg_day, nem_mcal_day, neg_mcal_day, cp_kg_day, ca_kg_day, p_kg_day,
  fodder_percent, concentrate_percent, source_sheet, is_active
) VALUES (
  {sql_str(p['id'])}, {sql_str(p['profile_code'])}, {sql_str(p['species'])},
  {sql_str(p['production_system'])}, {sql_str(p['life_stage'])},
  {sql_str(p['animal_class'])}, {sql_str(p.get('feed_cycle'))},
  {sql_num(p['body_weight_kg'])}, {sql_int(p['months_since_calving'])},
  {sql_num(p['dmi_kg_day'])},
  {sql_num(p['cp_percent_dm'])}, {sql_num(p['nel_mcal_per_kg_dm'])},
  {sql_num(p['ndf_percent_dm'])}, {sql_num(p['adf_percent_dm'])},
  {sql_num(p['ca_percent_dm'])}, {sql_num(p['p_percent_dm'])},
  {sql_num(p['milk_kg_day'])},
  {sql_num(p['tdn_kg_day'])}, {sql_num(p['nem_mcal_day'])}, {sql_num(p['neg_mcal_day'])},
  {sql_num(p['cp_kg_day'])}, {sql_num(p['ca_kg_day'])}, {sql_num(p['p_kg_day'])},
  {sql_num(p['fodder_percent'])}, {sql_num(p['concentrate_percent'])},
  {sql_str(p['source_sheet'])}, TRUE
) ON CONFLICT (profile_code) DO UPDATE SET
  dmi_kg_day = EXCLUDED.dmi_kg_day,
  cp_percent_dm = EXCLUDED.cp_percent_dm,
  nel_mcal_per_kg_dm = EXCLUDED.nel_mcal_per_kg_dm,
  ndf_percent_dm = EXCLUDED.ndf_percent_dm,
  adf_percent_dm = EXCLUDED.adf_percent_dm,
  ca_percent_dm = EXCLUDED.ca_percent_dm,
  p_percent_dm = EXCLUDED.p_percent_dm,
  tdn_kg_day = EXCLUDED.tdn_kg_day,
  nem_mcal_day = EXCLUDED.nem_mcal_day,
  cp_kg_day = EXCLUDED.cp_kg_day,
  ca_kg_day = EXCLUDED.ca_kg_day,
  p_kg_day = EXCLUDED.p_kg_day,
  feed_cycle = EXCLUDED.feed_cycle,
  updated_at = NOW();"""
        )
    path.write_text('\n'.join(lines) + '\n', encoding='utf-8')


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument('xlsx', nargs='?', default=str(DEFAULT_XLSX))
    args = parser.parse_args()
    xlsx = Path(args.xlsx)
    if not xlsx.exists():
        xlsx = FALLBACK_XLSX
    if not xlsx.exists():
        raise SystemExit(f'XLSX not found: {xlsx}')

    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    profiles = parse_dairy(wb) + parse_beef(wb) + small_ruminant_profiles()
    wb.close()
    profiles.extend(derived_extension_profiles(profiles))
    attach_feed_cycle(profiles)

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    MOBILE_ASSETS_DIR.mkdir(parents=True, exist_ok=True)

    json_path = MOBILE_ASSETS_DIR / 'nutrition_requirements.json'
    json_path.write_text(
        json.dumps({'profiles': profiles}, indent=2) + '\n',
        encoding='utf-8',
    )
    write_sql(profiles, OUT_DIR / 'nutrition_requirements.sql')
    print(f'Wrote {len(profiles)} profiles → {OUT_DIR / "nutrition_requirements.sql"}')
    print(f'JSON → {json_path}')


if __name__ == '__main__':
    main()
