#!/usr/bin/env python3
"""Extract breeds + weight curves from Excel workbooks into SQL seeds."""

from __future__ import annotations

import argparse
import json
import re
import uuid
from pathlib import Path

import openpyxl

SCRIPT_DIR = Path(__file__).resolve().parent
DB_DIR = SCRIPT_DIR.parent
OUT_DIR = DB_DIR / 'seeds'
SOURCES_DIR = DB_DIR / 'sources'
MOBILE_ASSETS_DIR = DB_DIR.parent.parent / 'assets' / 'data'

DEFAULT_BREEDS_XLSX = (
    Path.home() / 'Documents/Obsidian/GreenerHerd/Breed Information.xlsx'
)
DEFAULT_SHEEP_CHARS_XLSX = SOURCES_DIR / 'Sheep_Characteristics.xlsx'
FALLBACK_SHEEP_CHARS_XLSX = (
    Path.home() / 'Downloads/Sheep_Characteristics.xlsx'
)

NS = uuid.UUID('6ba7b810-9dad-11d1-80b4-00c04fd430c8')
AGE_MONTHS = (1, 4, 9, 15, 21, 24)


def slug_code(name: str) -> str:
    s = name.strip()
    s = re.sub(r'[/\\]+', '_', s)
    s = re.sub(r'[^A-Za-z0-9]+', '_', s)
    s = re.sub(r'_+', '_', s).strip('_').upper()
    return s[:80] or 'UNKNOWN'


def breed_id(species: str, code: str) -> str:
    return str(uuid.uuid5(NS, f'{species}:{code}'))


def norm_key(name: str) -> str:
    s = name.lower().replace('/', ' ').replace('-', ' ')
    s = re.sub(r'[^a-z0-9 ]+', '', s)
    return re.sub(r'\s+', ' ', s).strip()


def sql_str(v: object | None) -> str:
    if v is None:
        return 'NULL'
    s = str(v).replace("'", "''")
    return f"'{s}'"


def fix_longevity(val: object | None) -> str | None:
    if val is None:
        return None
    if isinstance(val, (int, float)) and val > 100:
        return None
    return str(val).strip() or None


def parse_weights(ws) -> dict[str, dict]:
    breeds: dict[str, dict] = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0 or not row or not row[0]:
            continue
        name_en = str(row[0]).strip()
        if name_en.lower() == 'breed':
            continue
        name_ar = str(row[1]).strip() if row[1] else None
        gender = str(row[2]).strip().upper() if row[2] else ''
        if gender not in ('MALE', 'FEMALE'):
            continue
        weights: dict[int, float] = {}
        for j, age in enumerate(AGE_MONTHS):
            val = row[3 + j]
            if val is not None:
                try:
                    weights[age] = float(val)
                except (TypeError, ValueError):
                    pass
        key = norm_key(name_en)
        if key not in breeds:
            breeds[key] = {
                'name_en': name_en,
                'name_ar': name_ar,
                'weights': {'MALE': {}, 'FEMALE': {}},
            }
        breeds[key]['weights'][gender].update(weights)
        if name_ar and not breeds[key].get('name_ar'):
            breeds[key]['name_ar'] = name_ar
    return breeds


def parse_characteristics_sheet(ws) -> dict[str, dict]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}
    headers = [str(h).strip() if h else '' for h in rows[0]]
    out: dict[str, dict] = {}
    for row in rows[1:]:
        if not row or not row[0]:
            continue
        name = str(row[0]).strip()
        if name.lower() == 'breed':
            continue
        rec = {headers[j]: row[j] for j in range(len(headers)) if headers[j]}
        out[norm_key(name)] = rec
        if '/' in name:
            for part in name.split('/'):
                p = part.strip()
                if p:
                    out[norm_key(p)] = rec
    return out


def load_sheep_characteristics(path: Path) -> dict[str, dict]:
    if not path.exists():
        raise FileNotFoundError(
            f'Sheep characteristics workbook not found: {path}\n'
            f'Copy Sheep_Characteristics.xlsx to {DEFAULT_SHEEP_CHARS_XLSX}',
        )
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = (
            wb['Sheep Characteristics']
            if 'Sheep Characteristics' in wb.sheetnames
            else wb[wb.sheetnames[0]]
        )
        chars = parse_characteristics_sheet(sheet)
        print(f'Loaded {len(chars)} sheep characteristic rows from {path.name}')
        return chars
    finally:
        wb.close()


def find_characteristics(key: str, chars: dict[str, dict]) -> dict:
    if key in chars:
        return chars[key]
    if 'shami' in key and 'damascus' in key:
        return chars.get(norm_key('Shami')) or chars.get(norm_key('Damascus')) or {}
    if 'naimi' in key and 'awassi' in key:
        return chars.get(norm_key('Naimi/Awassi')) or chars.get(norm_key('Awassi')) or {}
    if 'sawakini' in key or 'suakni' in key:
        return (
            chars.get(norm_key('Sawakini/Suakni'))
            or chars.get(norm_key('Sawakini'))
            or {}
        )
    for k, v in chars.items():
        if k == key or k in key or key in k:
            return v
    key_ascii = key.replace('é', 'e').replace('è', 'e')
    for k, v in chars.items():
        if k.replace('é', 'e').replace('è', 'e') == key_ascii:
            return v
    return {}


def char_field(c: dict, *names: str) -> object | None:
    for n in names:
        if n in c and c[n] is not None:
            return c[n]
    return None


def resolve_sheep_chars_path(explicit: Path | None) -> Path:
    if explicit is not None:
        return explicit
    if DEFAULT_SHEEP_CHARS_XLSX.exists():
        return DEFAULT_SHEEP_CHARS_XLSX
    if FALLBACK_SHEEP_CHARS_XLSX.exists():
        return FALLBACK_SHEEP_CHARS_XLSX
    return DEFAULT_SHEEP_CHARS_XLSX


def main() -> int:
    parser = argparse.ArgumentParser(
        description='Generate breeds.sql and breed_weights.sql from Excel sources',
    )
    parser.add_argument(
        'breeds_xlsx',
        nargs='?',
        type=Path,
        default=DEFAULT_BREEDS_XLSX,
        help='Main Breed Information.xlsx (weights + cattle/goat characteristics)',
    )
    parser.add_argument(
        '--sheep-chars',
        type=Path,
        default=None,
        help='Sheep_Characteristics.xlsx (default: services/db/sources/...)',
    )
    args = parser.parse_args()

    breeds_xlsx: Path = args.breeds_xlsx
    sheep_chars_path = resolve_sheep_chars_path(args.sheep_chars)

    if not breeds_xlsx.exists():
        print(f'File not found: {breeds_xlsx}')
        return 1

    wb = openpyxl.load_workbook(breeds_xlsx, read_only=True, data_only=True)
    cattle_w = parse_weights(wb['Cattle Weights'])
    goat_w = parse_weights(wb['Goat Weights old'])
    sheep_w = parse_weights(wb['Sheep Weights old'])
    cattle_c = parse_characteristics_sheet(wb['Cattle Characteristics'])
    goat_c = parse_characteristics_sheet(wb['Goat Charactieristics'])
    wb.close()

    sheep_c = load_sheep_characteristics(sheep_chars_path)

    breeds_sql: list[str] = [
        '-- Generated from Breed Information.xlsx + Sheep_Characteristics.xlsx',
        '-- Run: npm run db:extract-breeds (from services/db)',
        '',
        'DELETE FROM breed_weight_by_age;',
        'DELETE FROM breeds;',
        '',
    ]
    weights_sql: list[str] = [
        '-- Weight curves by breed, sex, and age (months)',
        '',
    ]

    total = 0
    weight_rows = 0
    sheep_with_chars = 0
    json_breeds: list[dict] = []

    for species, weights, chars in [
        ('CATTLE', cattle_w, cattle_c),
        ('GOAT', goat_w, goat_c),
        ('SHEEP', sheep_w, sheep_c),
    ]:
        for key in sorted(weights.keys(), key=lambda k: weights[k]['name_en']):
            w = weights[key]
            if w['name_en'].lower() == 'other':
                continue
            c = find_characteristics(key, chars)
            if species == 'SHEEP' and c:
                sheep_with_chars += 1

            code = slug_code(w['name_en'])
            bid = breed_id(species, code)

            birth_ease = char_field(
                c,
                'Calving Ease',
                'Kidding Ease',
                'Lambing Ease',
            )
            male_w = char_field(
                c,
                'Adult Bull Weight (kg)',
                'Adult Buck Weight (kg)',
                'Adult Ram Weight (kg)',
            )
            female_w = char_field(
                c,
                'Adult Cow Weight (kg)',
                'Adult Doe Weight (kg)',
                'Adult Ewe Weight (kg)',
            )

            json_breeds.append(
                {
                    'id': bid,
                    'species': species,
                    'code': code,
                    'name_en': w['name_en'],
                    'name_ar': w.get('name_ar'),
                    'origin': char_field(c, 'Origin'),
                    'primary_purpose': char_field(c, 'Primary Purpose'),
                    'color': char_field(c, 'Color'),
                    'milk_production_kg_year': char_field(c, 'Milk Production (kg/year)'),
                    'heat_tolerance': char_field(c, 'Heat Tolerance'),
                    'disease_resistance': char_field(c, 'Disease Resistance'),
                    'birth_ease': birth_ease,
                    'known_for': char_field(c, 'Known For'),
                }
            )

            breeds_sql.append(
                'INSERT INTO breeds (\n'
                '  id, species, code, name_en, name_ar, is_active,\n'
                '  origin, primary_purpose, color, milk_production_kg_year,\n'
                '  heat_tolerance, disease_resistance, birth_ease, feed_efficiency,\n'
                '  temperament, adult_male_weight_kg, adult_female_weight_kg,\n'
                '  adaptability, longevity_years, height_male_cm, height_female_cm, known_for\n'
                ') VALUES (\n'
                f"  '{bid}', '{species}', {sql_str(code)}, {sql_str(w['name_en'])}, "
                f"{sql_str(w.get('name_ar'))}, TRUE,\n"
                f"  {sql_str(char_field(c, 'Origin'))}, "
                f"{sql_str(char_field(c, 'Primary Purpose'))}, "
                f"{sql_str(char_field(c, 'Color'))}, "
                f"{sql_str(char_field(c, 'Milk Production (kg/year)'))},\n"
                f"  {sql_str(char_field(c, 'Heat Tolerance'))}, "
                f"{sql_str(char_field(c, 'Disease Resistance'))}, "
                f"{sql_str(birth_ease)}, "
                f"{sql_str(char_field(c, 'Feed Efficiency'))},\n"
                f"  {sql_str(char_field(c, 'Temperament'))}, "
                f"{sql_str(male_w)}, "
                f"{sql_str(female_w)},\n"
                f"  {sql_str(char_field(c, 'Adaptability'))}, "
                f"{sql_str(fix_longevity(char_field(c, 'Longevity (years)')))}, "
                f"{sql_str(char_field(c, 'Height-Male (cm)'))}, "
                f"{sql_str(char_field(c, 'Height-Female (cm)'))}, "
                f"{sql_str(char_field(c, 'Known For'))}\n"
                ');'
            )
            total += 1

            for sex in ('MALE', 'FEMALE'):
                for age, kg in sorted(w['weights'][sex].items()):
                    weights_sql.append(
                        'INSERT INTO breed_weight_by_age (breed_id, sex, age_months, weight_kg) VALUES '
                        f"('{bid}', '{sex}', {age}, {kg});"
                    )
                    weight_rows += 1

    breeds_path = OUT_DIR / 'breeds.sql'
    weights_path = OUT_DIR / 'breed_weights.sql'
    breeds_path.write_text('\n'.join(breeds_sql) + '\n', encoding='utf-8')
    weights_path.write_text('\n'.join(weights_sql) + '\n', encoding='utf-8')

    MOBILE_ASSETS_DIR.mkdir(parents=True, exist_ok=True)
    json_path = MOBILE_ASSETS_DIR / 'breeds.json'
    json_path.write_text(
        json.dumps({'breeds': json_breeds}, ensure_ascii=False, indent=2) + '\n',
        encoding='utf-8',
    )

    print(f'Wrote {total} breeds -> {breeds_path}')
    print(f'Wrote mobile catalog -> {json_path}')
    print(f'Wrote {weight_rows} weight rows -> {weights_path}')
    print(f'Sheep breeds with characteristics: {sheep_with_chars}/{len(sheep_w) - (1 if "other" in sheep_w else 0)}')
    by_species = {
        'CATTLE': sum(1 for k in cattle_w if cattle_w[k]['name_en'].lower() != 'other'),
        'GOAT': sum(1 for k in goat_w if goat_w[k]['name_en'].lower() != 'other'),
        'SHEEP': sum(1 for k in sheep_w if sheep_w[k]['name_en'].lower() != 'other'),
    }
    print('Per species:', by_species)
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
