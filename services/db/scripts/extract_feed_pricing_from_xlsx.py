#!/usr/bin/env python3
"""Extract indicative pricing + FX rates from Excel into SQL seeds + JSON."""

from __future__ import annotations

import argparse
import json
import uuid
from pathlib import Path

import openpyxl

SCRIPT_DIR = Path(__file__).resolve().parent
DB_DIR = SCRIPT_DIR.parent
OUT_DIR = DB_DIR / 'seeds'
SOURCES_DIR = DB_DIR / 'sources'
MOBILE_ASSETS_DIR = DB_DIR.parent.parent / 'assets' / 'data'
DEFAULT_XLSX = SOURCES_DIR / 'Generic_Products_Indicative_Pricing.xlsx'
FALLBACK_XLSX = Path.home() / 'Downloads/Generic_Products_Indicative_Pricing.xlsx'

NS = uuid.UUID('a3b5c7d9-e1f2-4a6b-8c0d-111111111112')

TYPE_MAP = {
    'fodder': 'FODDER',
    'concentrate': 'CONCENTRATE',
    'additive': 'ADDITIVE',
}

# Excel column header -> (ISO country code, currency code)
COUNTRY_COLUMNS: list[tuple[str, str, str]] = [
    ('UAE (AED/kg)', 'AE', 'AED'),
    ('Bahrain (BHD/kg)', 'BH', 'BHD'),
    ('Kuwait (KWD/kg)', 'KW', 'KWD'),
    ('Qatar (QAR/kg)', 'QA', 'QAR'),
    ('Saudi Arabia (SAR/kg)', 'SA', 'SAR'),
    ('Oman (OMR/kg)', 'OM', 'OMR'),
    ('Jordan (JOD/kg)', 'JO', 'JOD'),
    ('Palestine (ILS/kg)', 'PS', 'ILS'),
    ('Israel (ILS/kg)', 'IL', 'ILS'),
    ('Iraq (IQD/kg)', 'IQ', 'IQD'),
    ('Egypt (EGP/kg)', 'EG', 'EGP'),
    ('Syria (SYP/kg)', 'SY', 'SYP'),
    ('Lebanon (USD/kg)', 'LB', 'USD'),
    ('Libya (USD/kg)', 'LY', 'USD'),
    ('Algeria (DZD/kg)', 'DZ', 'DZD'),
    ('Morocco (MAD/kg)', 'MA', 'MAD'),
    ('Tunisia (TND/kg)', 'TN', 'TND'),
    ('Yemen (YER/kg)', 'YE', 'YER'),
    ('Sudan (SDG/kg)', 'SD', 'SDG'),
]


def sql_str(v: object | None) -> str:
    if v is None:
        return 'NULL'
    return "'" + str(v).replace("'", "''") + "'"


def sql_num(v: object | None) -> str:
    if v is None or v == '':
        return 'NULL'
    try:
        return str(float(v))
    except (TypeError, ValueError):
        return 'NULL'


def map_type(raw: object | None) -> str:
    key = (str(raw).strip().lower() if raw else 'fodder')
    return TYPE_MAP.get(key, 'FODDER')


def price_id(name_en: str, feed_type: str) -> str:
    return str(uuid.uuid5(NS, f'indicative-price:{name_en}:{feed_type}'))


def clean_text(v: object | None) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument('xlsx', nargs='?', default=str(DEFAULT_XLSX))
    args = parser.parse_args()
    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        xlsx_path = FALLBACK_XLSX
    if not xlsx_path.exists():
        raise SystemExit(f'Workbook not found: {args.xlsx}')

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    # FX rates
    fx_ws = wb['FX Rates Reference']
    fx_rows = list(fx_ws.iter_rows(values_only=True))
    fx_headers = {str(h).strip(): i for i, h in enumerate(fx_rows[0]) if h}
    fx_records: list[dict] = []
    country_to_code = {
        'UAE': 'AE',
        'Bahrain': 'BH',
        'Kuwait': 'KW',
        'Qatar': 'QA',
        'Saudi Arabia': 'SA',
        'Oman': 'OM',
        'Jordan': 'JO',
        'Palestine': 'PS',
        'Israel': 'IL',
        'Iraq': 'IQ',
        'Egypt': 'EG',
        'Syria': 'SY',
        'Lebanon': 'LB',
        'Libya': 'LY',
        'Algeria': 'DZ',
        'Morocco': 'MA',
        'Tunisia': 'TN',
        'Yemen': 'YE',
        'Sudan': 'SD',
    }
    for row in fx_rows[1:]:
        if not row or not row[fx_headers['Country']]:
            continue
        country_name = str(row[fx_headers['Country']]).strip()
        code = country_to_code.get(country_name)
        if not code:
            continue
        fx_records.append(
            {
                'country_code': code,
                'country_name': country_name,
                'currency_code': str(row[fx_headers['Code']]).strip(),
                'rate_per_usd': float(row[fx_headers['Rate per USD']]),
                'peg_type': clean_text(row[fx_headers.get('Peg Type')]),
                'notes': clean_text(row[fx_headers.get('Notes')]),
            }
        )

    # Indicative prices
    ws = wb['Indicative Pricing']
    rows = list(ws.iter_rows(values_only=True))
    headers = {str(h).strip(): i for i, h in enumerate(rows[0]) if h}

    price_records: list[dict] = []
    price_number = 2001
    for row in rows[1:]:
        if not row or not row[headers['Product Name']]:
            continue
        name_en = str(row[headers['Product Name']]).strip()
        feed_type = map_type(row[headers['Type']])
        prices_by_country: dict[str, dict] = {}
        for col_name, country_code, currency_code in COUNTRY_COLUMNS:
            if col_name not in headers:
                continue
            val = row[headers[col_name]]
            if val is None or val == '':
                continue
            prices_by_country[country_code] = {
                'currency': currency_code,
                'price_per_kg': round(float(val), 6),
            }

        price_records.append(
            {
                'id': price_id(name_en, feed_type),
                'price_number': price_number,
                'name_en': name_en,
                'name_ar': clean_text(row[headers['Arabic Name']]),
                'feed_type': feed_type,
                'usd_per_kg': row[headers['USD Base ($/kg)']],
                'prices_by_country': prices_by_country,
                'price_source': clean_text(row[headers['Price Source']]),
            }
        )
        price_number += 1

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    MOBILE_ASSETS_DIR.mkdir(parents=True, exist_ok=True)

    fx_sql = [
        '-- Generated by extract_feed_pricing_from_xlsx.py',
        'TRUNCATE feed_fx_rates CASCADE;',
        'INSERT INTO feed_fx_rates (country_code, country_name, currency_code, rate_per_usd, peg_type, notes)',
        'VALUES',
    ]
    fx_vals = [
        '('
        + ', '.join(
            [
                sql_str(r['country_code']),
                sql_str(r['country_name']),
                sql_str(r['currency_code']),
                sql_num(r['rate_per_usd']),
                sql_str(r['peg_type']),
                sql_str(r['notes']),
            ]
        )
        + ')'
        for r in fx_records
    ]
    (OUT_DIR / 'feed_fx_rates.sql').write_text(
        '\n'.join(fx_sql + [',\n'.join(fx_vals) + ';']),
        encoding='utf-8',
    )

    price_sql = [
        '-- Generated by extract_feed_pricing_from_xlsx.py',
        'TRUNCATE feed_indicative_prices RESTART IDENTITY CASCADE;',
        'INSERT INTO feed_indicative_prices (',
        '  id, price_number, name_en, name_ar, feed_type, usd_per_kg,',
        '  prices_by_country, price_source, is_active',
        ') VALUES',
    ]
    price_vals = []
    for r in price_records:
        price_vals.append(
            '('
            + ', '.join(
                [
                    f"'{r['id']}'::uuid",
                    str(r['price_number']),
                    sql_str(r['name_en']),
                    sql_str(r['name_ar']),
                    sql_str(r['feed_type']),
                    sql_num(r['usd_per_kg']),
                    sql_str(json.dumps(r['prices_by_country'], ensure_ascii=False))
                    + '::jsonb',
                    sql_str(r['price_source']),
                    'TRUE',
                ]
            )
            + ')'
        )
    (OUT_DIR / 'feed_indicative_prices.sql').write_text(
        '\n'.join(price_sql + [',\n'.join(price_vals) + ';']),
        encoding='utf-8',
    )

    json_path = MOBILE_ASSETS_DIR / 'feed_indicative_prices.json'
    json_path.write_text(
        json.dumps(
            {
                'prices': price_records,
                'fx_rates': fx_records,
                'meta': {
                    'price_count': len(price_records),
                    'fx_count': len(fx_records),
                    'price_number_start': price_records[0]['price_number'],
                    'price_number_end': price_records[-1]['price_number'],
                },
            },
            indent=2,
            ensure_ascii=False,
        ),
        encoding='utf-8',
    )

    print(f'Wrote {len(fx_records)} FX rates -> {OUT_DIR / "feed_fx_rates.sql"}')
    print(
        f'Wrote {len(price_records)} indicative prices -> {OUT_DIR / "feed_indicative_prices.sql"}'
    )
    print(f'Wrote mobile JSON -> {json_path}')


if __name__ == '__main__':
    main()
