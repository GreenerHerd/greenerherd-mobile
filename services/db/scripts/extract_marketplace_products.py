#!/usr/bin/env python3
"""Extract marketplace feed products into SQL + mobile JSON.

Sources (first match wins):
  1. sources/Marketplace_Feed_Products.xlsx  (preferred — your marketplace sheet)
  2. utility-feed-product-agent-genui/db/seed_data_market_products*.sql (fallback)

Nutrition fields are joined from feed_products.json when product names match.
"""

from __future__ import annotations

import argparse
import json
import re
import uuid
from pathlib import Path

import openpyxl

SCRIPT_DIR = Path(__file__).resolve().parent
DB_DIR = SCRIPT_DIR.parent
OUT_DIR = DB_DIR / 'seeds'
SOURCES_DIR = DB_DIR / 'sources'
MOBILE_ASSETS_DIR = DB_DIR.parent.parent / 'assets' / 'data'
DEFAULT_XLSX = SOURCES_DIR / 'Marketplace_Feed_Products.xlsx'
# Repo layout: Greenerherd/greenerherd-mobile/services/db  +  Greenerherd/utility-feed-product-agent-genui/db
_GREENERHERD_ROOT = DB_DIR.parent.parent.parent
UTILITY_DB_DIR = _GREENERHERD_ROOT / 'utility-feed-product-agent-genui' / 'db'
FEED_PRODUCTS_JSON = MOBILE_ASSETS_DIR / 'feed_products.json'

NS = uuid.UUID('8f4e3c21-9b2a-4f6d-a1e0-000000000002')
PROD_ID_START = 3001

TYPE_MAP = {
    'fodder': 'FODDER',
    'concentrate': 'CONCENTRATE',
    'additive': 'ADDITIVE',
}

COUNTRY_TO_CODE = {
    'uae': 'AE',
    'united arab emirates': 'AE',
    'saudi arabia': 'SA',
    'ksa': 'SA',
    'saudi': 'SA',
    'qatar': 'QA',
    'bahrain': 'BH',
    'kuwait': 'KW',
    'oman': 'OM',
    'jordan': 'JO',
    'egypt': 'EG',
    'morocco': 'MA',
    'tunisia': 'TN',
    'lebanon': 'LB',
    'iraq': 'IQ',
    'palestine': 'PS',
    'israel': 'IL',
    'algeria': 'DZ',
    'yemen': 'YE',
    'sudan': 'SD',
    'libya': 'LY',
    'syria': 'SY',
}


def sql_str(v: object | None) -> str:
    if v is None:
        return 'NULL'
    return "'" + str(v).replace("'", "''") + "'"


def sql_num(v: object | None) -> str:
    if v is None or v == '':
        return 'NULL'
    try:
        return str(float(v))
    except (TypeError, ValueError):
        return 'NULL'


def sql_int(v: object | None) -> str:
    if v is None or v == '':
        return 'NULL'
    try:
        return str(int(float(v)))
    except (TypeError, ValueError):
        return 'NULL'


def clean_text(v: object | None) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def map_type(raw: object | None) -> str:
    key = (str(raw).strip().lower() if raw else 'fodder')
    return TYPE_MAP.get(key, 'FODDER')


def map_country(raw: object | None) -> str:
    if raw is None:
        return 'SA'
    s = str(raw).strip()
    if len(s) == 2:
        return s.upper()
    return COUNTRY_TO_CODE.get(s.lower(), 'SA')


def product_uuid(marketplace_product_id: str) -> str:
    return str(uuid.uuid5(NS, f'marketplace:{marketplace_product_id}'))


def normalize_name(name: str) -> str:
    return re.sub(r'[^a-z0-9]+', '', name.lower())


MARKETPLACE_NAME_ALIASES: dict[str, str] = {
    'soyabeanmeal': 'soybeanmeal',
}

# Marketing / regional tokens stripped before fuzzy match.
_STRIP_TOKENS = frozenset({
    'premium',
    'organic',
    'good',
    'quality',
    'riyadh',
    'jeddah',
})


def match_standard_product(
    name_en: str, nutrition_index: dict[str, dict]
) -> dict | None:
    """Resolve a marketplace listing name to a standard feed_products row."""
    key = normalize_name(name_en)
    key = MARKETPLACE_NAME_ALIASES.get(key, key)
    if key in nutrition_index:
        return nutrition_index[key]

    best: dict | None = None
    best_len = 0
    for cat_key, product in nutrition_index.items():
        if len(cat_key) < 4:
            continue
        if cat_key in key or key in cat_key:
            if len(cat_key) > best_len:
                best_len = len(cat_key)
                best = product
    if best is not None:
        return best

    tokens = [
        t
        for t in re.split(r'[^a-z0-9]+', name_en.lower())
        if len(t) > 2 and t not in _STRIP_TOKENS
    ]
    if not tokens:
        return None
    for product in nutrition_index.values():
        name_tokens = [
            t
            for t in re.split(r'[^a-z0-9]+', product.get('name_en', '').lower())
            if len(t) > 2
        ]
        if all(
            any(ct in token or token in ct for ct in name_tokens)
            for token in tokens[:2]
        ):
            return product
    return None


def load_nutrition_index() -> dict[str, dict]:
    if not FEED_PRODUCTS_JSON.exists():
        return {}
    data = json.loads(FEED_PRODUCTS_JSON.read_text(encoding='utf-8'))
    index: dict[str, dict] = {}
    for p in data.get('products', []):
        key = normalize_name(p.get('name_en', ''))
        if key:
            index[key] = p
    return index


def nutrition_for(name_en: str, nutrition_index: dict[str, dict]) -> dict:
    p = match_standard_product(name_en, nutrition_index)
    if not p:
        return {}
    return {
        'standard_product_number': p.get('product_number'),
        'eligibility_rules': p.get('eligibility_rules') or [],
        'dm_percent': p.get('dm_percent'),
        'cp_percent': p.get('cp_percent'),
        'nem_mcal_kg': p.get('nem_mcal_kg'),
        'ndf_percent': p.get('ndf_percent'),
    }


def parse_utility_sql() -> list[dict]:
    files = [
        UTILITY_DB_DIR / 'seed_data_market_products.sql',
        UTILITY_DB_DIR / 'seed_data_market_products_extended.sql',
    ]
    pat = re.compile(
        r"\('((?:[^']|'')*)',\s*'(MP-[^']+)',\s*'((?:[^']|'')*)',\s*'"
        r"(Fodder|Concentrate|Additive)',\s*([\d.]+),\s*'([A-Z]{3})',\s*'"
        r"((?:[^']|'')*)',\s*'([^']+)',\s*'((?:[^']|'')*)',\s*'"
        r"((?:[^']|'')*)',\s*'((?:[^']|'')*)',\s*false,\s*(\d+),\s*true\)",
        re.I,
    )
    records: list[dict] = []
    seen_codes: set[str] = set()
    for path in files:
        if not path.exists():
            continue
        text = path.read_text(encoding='utf-8', errors='ignore')
        for m in pat.finditer(text):
            code = m.group(2)
            if code in seen_codes:
                continue
            seen_codes.add(code)
            records.append(
                {
                    'marketplace_product_id': code,
                    'name_en': m.group(1).replace("''", "'"),
                    'name_ar': None,
                    'feed_type': map_type(m.group(4)),
                    'price_per_kg': float(m.group(5)),
                    'currency': m.group(6),
                    'supplier_name': m.group(7).replace("''", "'"),
                    'supplier_country': m.group(8),
                    'supplier_email': clean_text(m.group(9).replace("''", "'")),
                    'supplier_phone': clean_text(m.group(10).replace("''", "'")),
                    'supplier_address': clean_text(m.group(11).replace("''", "'")),
                    'source': 'Marketplace seed (utility SQL)',
                }
            )
    return records


def parse_xlsx(path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet_name = None
    for name in wb.sheetnames:
        if 'market' in name.lower() or 'product' in name.lower():
            sheet_name = name
            break
    ws = wb[sheet_name or wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = {str(h).strip().lower(): i for i, h in enumerate(rows[0]) if h}

    def col(*names: str) -> int | None:
        for n in names:
            if n in headers:
                return headers[n]
        return None

    i_name = col('product name', 'name_en', 'name')
    i_ar = col('arabic name', 'name_ar')
    i_code = col('product code', 'marketplace_product_id', 'product_code', 'code')
    i_type = col('type', 'feed_type', 'category')
    i_price = col('price_per_kg', 'cost_per_kg', 'price/kg', 'price')
    i_currency = col('currency', 'cost_currency')
    i_supplier = col('supplier', 'supplier_name')
    i_country = col('supplier_country', 'country', 'country_code')
    i_email = col('supplier_email', 'email')
    i_phone = col('supplier_phone', 'phone')
    i_address = col('supplier_address', 'address')
    i_min = col('min_order_kg', 'min order kg')
    i_pack = col('pack_size_kg', 'pack size kg')

    if i_name is None or i_supplier is None:
        raise SystemExit(
            'XLSX must include at least Product Name and Supplier columns. '
            f'Found headers: {list(headers.keys())}'
        )

    records: list[dict] = []
    seq = 0
    for row in rows[1:]:
        if not row or i_name >= len(row) or not row[i_name]:
            continue
        seq += 1
        name_en = str(row[i_name]).strip()
        code = (
            str(row[i_code]).strip()
            if i_code is not None and i_code < len(row) and row[i_code]
            else f'MP-{seq:04d}'
        )
        country_raw = (
            row[i_country] if i_country is not None and i_country < len(row) else 'SA'
        )
        records.append(
            {
                'marketplace_product_id': code,
                'name_en': name_en,
                'name_ar': clean_text(row[i_ar]) if i_ar is not None and i_ar < len(row) else None,
                'feed_type': map_type(row[i_type] if i_type is not None and i_type < len(row) else None),
                'price_per_kg': float(row[i_price]) if i_price is not None and i_price < len(row) and row[i_price] else 0,
                'currency': (
                    str(row[i_currency]).strip()
                    if i_currency is not None and i_currency < len(row) and row[i_currency]
                    else 'SAR'
                ),
                'supplier_name': str(row[i_supplier]).strip(),
                'supplier_country': country_raw,
                'supplier_email': clean_text(row[i_email]) if i_email is not None and i_email < len(row) else None,
                'supplier_phone': clean_text(row[i_phone]) if i_phone is not None and i_phone < len(row) else None,
                'supplier_address': clean_text(row[i_address]) if i_address is not None and i_address < len(row) else None,
                'min_order_kg': row[i_min] if i_min is not None and i_min < len(row) else None,
                'pack_size_kg': row[i_pack] if i_pack is not None and i_pack < len(row) else None,
                'source': 'Marketplace_Feed_Products.xlsx',
            }
        )
    return records


def finalize_records(raw: list[dict], nutrition_index: dict[str, dict]) -> list[dict]:
    records: list[dict] = []
    prod_id = PROD_ID_START
    for r in sorted(raw, key=lambda x: x['marketplace_product_id']):
        mp_id = r['marketplace_product_id']
        extra = nutrition_for(r['name_en'], nutrition_index)
        records.append(
            {
                'id': product_uuid(mp_id),
                'prod_id': prod_id,
                'marketplace_product_id': mp_id,
                'name_en': r['name_en'],
                'name_ar': r.get('name_ar'),
                'feed_type': r['feed_type'],
                'supplier_name': r['supplier_name'],
                'supplier_email': r.get('supplier_email'),
                'supplier_phone': r.get('supplier_phone'),
                'supplier_address': r.get('supplier_address'),
                'country_code': map_country(r.get('supplier_country')),
                'currency': (r.get('currency') or 'SAR')[:3].upper(),
                'price_per_kg': r['price_per_kg'],
                'min_order_kg': r.get('min_order_kg'),
                'pack_size_kg': r.get('pack_size_kg'),
                'standard_product_number': extra.get('standard_product_number'),
                'dm_percent': extra.get('dm_percent'),
                'cp_percent': extra.get('cp_percent'),
                'nem_mcal_kg': extra.get('nem_mcal_kg'),
                'ndf_percent': extra.get('ndf_percent'),
                'in_stock': True,
                'source': r.get('source'),
            }
        )
        prod_id += 1
    return records


def record_to_sql(r: dict) -> str:
    vals = [
        f"'{r['id']}'::uuid",
        str(r['prod_id']),
        sql_str(r['marketplace_product_id']),
        sql_str(r['name_en']),
        sql_str(r['name_ar']),
        sql_str(r['feed_type']),
        sql_str(r['supplier_name']),
        sql_str(r['supplier_email']),
        sql_str(r['supplier_phone']),
        sql_str(r['supplier_address']),
        sql_str(r['country_code']),
        sql_str(r['currency']),
        sql_num(r['price_per_kg']),
        sql_num(r['min_order_kg']),
        sql_num(r['pack_size_kg']),
        sql_int(r['standard_product_number']),
        sql_num(r['dm_percent']),
        sql_num(r['cp_percent']),
        sql_num(r['nem_mcal_kg']),
        sql_num(r['ndf_percent']),
        'TRUE',
        'TRUE',
        sql_str(r['source']),
    ]
    return f"({', '.join(vals)})"


def build_upload_sql(records: list[dict]) -> str:
    if not records:
        return '-- No marketplace products\n'
    start = records[0]['prod_id']
    end = records[-1]['prod_id']
    update_cols = [
        'name_en', 'name_ar', 'feed_type', 'supplier_name', 'supplier_email',
        'supplier_phone', 'supplier_address', 'country_code', 'currency',
        'price_per_kg', 'min_order_kg', 'pack_size_kg', 'standard_product_number',
        'dm_percent', 'cp_percent', 'nem_mcal_kg', 'ndf_percent',
        'in_stock', 'is_active', 'source', 'updated_at',
    ]
    set_clause = ',\n  '.join(
        f'{c} = EXCLUDED.{c}' if c != 'updated_at' else 'updated_at = NOW()'
        for c in update_cols
    )
    lines = [
        '-- Generated by extract_marketplace_products.py',
        '-- Upload marketplace feed listings',
        '--',
        '-- prod_id = marketplace_feed_products.prod_id (unique int, 3001+)',
        '-- marketplace_product_id = external listing code (e.g. MP-UAE-001-001)',
        f'-- Rows: {len(records)} products (prod_id {start}–{end})',
        '-- Safe to re-run: upserts on marketplace_product_id',
        '-- Prerequisite: migrations/012_marketplace_feed_products.sql',
        '',
        'BEGIN;',
        '',
        'INSERT INTO marketplace_feed_products (',
        '  id, prod_id, marketplace_product_id, name_en, name_ar, feed_type,',
        '  supplier_name, supplier_email, supplier_phone, supplier_address,',
        '  country_code, currency, price_per_kg, min_order_kg, pack_size_kg,',
        '  standard_product_number, dm_percent, cp_percent, nem_mcal_kg, ndf_percent,',
        '  in_stock, is_active, source',
        ') VALUES',
        ',\n'.join(record_to_sql(r) for r in records),
        'ON CONFLICT (marketplace_product_id) DO UPDATE SET',
        f'  {set_clause};',
        '',
        'INSERT INTO marketplace_feed_product_trans (product_id, locale, name)',
        'SELECT id, \'en\', name_en FROM marketplace_feed_products',
        f'WHERE prod_id BETWEEN {start} AND {end}',
        'ON CONFLICT (product_id, locale) DO UPDATE SET name = EXCLUDED.name;',
        '',
        'INSERT INTO marketplace_feed_product_trans (product_id, locale, name)',
        'SELECT id, \'ar\', name_ar FROM marketplace_feed_products',
        f'WHERE prod_id BETWEEN {start} AND {end} AND name_ar IS NOT NULL',
        'ON CONFLICT (product_id, locale) DO UPDATE SET name = EXCLUDED.name;',
        '',
        'COMMIT;',
        '',
    ]
    return '\n'.join(lines)


def mobile_json(records: list[dict]) -> dict:
    products = []
    for r in records:
        entry: dict = {
            'id': r['marketplace_product_id'],
            'name_en': r['name_en'],
            'name_ar': r['name_ar'] or '',
            'feed_type': r['feed_type'],
            'supplier_name': r['supplier_name'],
            'supplier_phone': r['supplier_phone'] or '',
            'country_code': r['country_code'],
            'currency': r['currency'],
            'price_per_kg': round(float(r['price_per_kg']), 4),
            'min_order_kg': r['min_order_kg'],
            'pack_size_kg': r['pack_size_kg'],
            'dm_percent': r['dm_percent'],
            'cp_percent': r['cp_percent'],
            'nem_mcal_kg': r['nem_mcal_kg'],
            'ndf_percent': r['ndf_percent'],
            'in_stock': r['in_stock'],
        }
        if r.get('standard_product_number') is not None:
            entry['standard_product_number'] = r['standard_product_number']
        rules = r.get('eligibility_rules') or []
        if rules:
            entry['eligibility_rules'] = rules
        products.append(entry)
    return {
        'products': products,
        'meta': {
            'total': len(products),
            'prod_id_start': records[0]['prod_id'] if records else None,
            'prod_id_end': records[-1]['prod_id'] if records else None,
        },
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument('xlsx', nargs='?', default=str(DEFAULT_XLSX))
    args = parser.parse_args()

    nutrition_index = load_nutrition_index()
    xlsx_path = Path(args.xlsx)
    if xlsx_path.exists():
        raw = parse_xlsx(xlsx_path)
        print(f'Loaded {len(raw)} rows from {xlsx_path}')
    else:
        raw = parse_utility_sql()
        print(
            f'XLSX not found at {xlsx_path}; '
            f'loaded {len(raw)} rows from utility SQL seeds'
        )

    if not raw:
        raise SystemExit('No marketplace products found')

    records = finalize_records(raw, nutrition_index)
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    MOBILE_ASSETS_DIR.mkdir(parents=True, exist_ok=True)

    cols = (
        '  id, prod_id, marketplace_product_id, name_en, name_ar, feed_type,\n'
        '  supplier_name, supplier_email, supplier_phone, supplier_address,\n'
        '  country_code, currency, price_per_kg, min_order_kg, pack_size_kg,\n'
        '  standard_product_number, dm_percent, cp_percent, nem_mcal_kg, ndf_percent,\n'
        '  in_stock, is_active, source'
    )
    seed_sql = [
        '-- Generated by extract_marketplace_products.py',
        'TRUNCATE marketplace_feed_product_trans CASCADE;',
        'TRUNCATE marketplace_feed_products RESTART IDENTITY CASCADE;',
        'INSERT INTO marketplace_feed_products (',
        cols,
        ') VALUES',
        ',\n'.join(record_to_sql(r) for r in records) + ';',
    ]
    seed_path = OUT_DIR / 'marketplace_feed_products.sql'
    seed_path.write_text('\n'.join(seed_sql), encoding='utf-8')

    upload_path = OUT_DIR / 'upload_marketplace_feed_products.sql'
    upload_path.write_text(build_upload_sql(records), encoding='utf-8')

    json_path = MOBILE_ASSETS_DIR / 'marketplace_feed_products.json'
    json_path.write_text(
        json.dumps(mobile_json(records), indent=2, ensure_ascii=False),
        encoding='utf-8',
    )

    by_country: dict[str, int] = {}
    by_type: dict[str, int] = {}
    for r in records:
        by_country[r['country_code']] = by_country.get(r['country_code'], 0) + 1
        by_type[r['feed_type']] = by_type.get(r['feed_type'], 0) + 1

    print(f'Wrote {len(records)} products -> {seed_path}')
    print(f'Wrote upload DML -> {upload_path}')
    print(f'Wrote mobile JSON -> {json_path}')
    print('By country_code:', dict(sorted(by_country.items())))
    print('By feed_type:', by_type)


if __name__ == '__main__':
    main()
